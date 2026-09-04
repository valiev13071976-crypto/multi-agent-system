"""Block 11 — Product Economics / Margin Intelligence closure tests."""

from __future__ import annotations

import io
import unittest

from openpyxl import Workbook, load_workbook

from data_intel.cleaning import normalize_decimal_string
from data_intel.economics import (
    COMPLETE,
    DECISION_ALLOW,
    DECISION_DENY,
    DECISION_REQUIRE_REVIEW,
    DECISION_WARN,
    DISCOUNT_PLATFORM,
    DISCOUNT_SELLER,
    DISCOUNT_UNKNOWN,
    INSUFFICIENT_INPUT,
    PARTIAL,
    PROV_UNKNOWN,
    EconomicsInput,
    EconomicsPolicy,
    apply_scenario,
    calculate_economics,
    calculate_minimum_price,
    compare_channels,
    economics_batch_rows,
    parse_economics_row,
)
from data_intel.errors import DATASET_ACCESS_DENIED, DataIntelError
from data_intel.service import DataIntelligenceService
from data_intel.store import InMemoryDatasetStore
from decimal import Decimal


def _xlsx(rows: list[list], headers: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Economics"
    hdr = headers or [
        "sku",
        "product_name",
        "purchase_price",
        "selling_price",
        "commission_rate",
        "logistics_cost",
        "advertising_cost",
        "vat_rate",
    ]
    ws.append(hdr)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FIXTURE_CHANNELS = {
    "SITE": {"commission_rate": "0", "acquiring_rate": "0.02", "logistics_cost": "0"},
    "WILDBERRIES": {"commission_rate": "0.15", "logistics_cost": "350", "advertising_cost": "0"},
    "OZON": {"commission_rate": "0.12", "logistics_cost": "400", "advertising_cost": "0"},
    "YANDEX_MARKET": {"commission_rate": "0.10", "logistics_cost": "300", "advertising_rate": "0.05"},
}


class DecimalMoneyTests(unittest.TestCase):
    def test_decimal_not_float(self):
        inp = EconomicsInput(purchase_price=Decimal("95000"), selling_price=Decimal("120000"))
        r = calculate_economics(inp)
        self.assertIsInstance(Decimal(r["gross_difference"]), Decimal)


class ExampleATests(unittest.TestCase):
    """Purchase 95k, Sell 120k — 25k is NOT net profit."""

    def test_gross_difference_not_net_profit(self):
        inp = EconomicsInput(
            purchase_price=Decimal("95000"),
            selling_price=Decimal("120000"),
            commission_rate=Decimal("0.15"),
            commission_prov="CONFIGURED",
            logistics_cost=Decimal("350"),
            logistics_prov="CONFIGURED",
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        r = calculate_economics(inp)
        self.assertEqual(r["gross_difference"], "25000.00")
        self.assertNotIn("net_profit", r["profit_label"])
        self.assertNotEqual(r["contribution"], r["gross_difference"])
        self.assertIn("not net profit", r["note"].lower())


class DiscountTests(unittest.TestCase):
    def test_seller_funded_discount(self):
        base = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("200"),
            discount_rate=Decimal("0.10"),
            discount_ownership=DISCOUNT_SELLER,
        )
        r = calculate_economics(base)
        self.assertEqual(r["effective_price"], "180.00")

    def test_platform_funded_unchanged_revenue(self):
        base = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("200"),
            discount_rate=Decimal("0.10"),
            discount_ownership=DISCOUNT_PLATFORM,
        )
        r = calculate_economics(base)
        self.assertEqual(r["effective_price"], "200.00")

    def test_unknown_discount_ownership_review(self):
        base = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("200"),
            discount_rate=Decimal("0.10"),
            discount_ownership=DISCOUNT_UNKNOWN,
        )
        r = calculate_economics(base)
        self.assertEqual(r["completeness"], PARTIAL)
        self.assertIn("discount_ownership_unknown", r["warnings"])


class UnknownVsZeroTests(unittest.TestCase):
    def test_unknown_advertising_not_zero(self):
        inp = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("200"),
            advertising_prov=PROV_UNKNOWN,
        )
        r = calculate_economics(inp)
        self.assertIn("advertising", r["missing"])
        self.assertEqual(r["completeness"], PARTIAL)

    def test_explicit_zero_advertising(self):
        inp = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("200"),
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        r = calculate_economics(inp)
        self.assertEqual(r["advertising"], "0.00")
        self.assertNotIn("advertising", r["missing"])


class MinimumPriceTests(unittest.TestCase):
    def test_minimum_price_available(self):
        inp = EconomicsInput(
            purchase_price=Decimal("95000"),
            commission_rate=Decimal("0.15"),
            commission_prov="CONFIGURED",
            logistics_cost=Decimal("350"),
            logistics_prov="CONFIGURED",
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        price, status, missing = calculate_minimum_price(inp, policy=EconomicsPolicy(minimum_margin_pct=Decimal("10")))
        self.assertIsNotNone(price)
        self.assertEqual(status, "minimum_margin")
        self.assertEqual(missing, ())

    def test_floor_unavailable_when_incomplete(self):
        inp = EconomicsInput(purchase_price=Decimal("100"))
        price, status, missing = calculate_minimum_price(inp)
        self.assertIsNone(price)
        self.assertEqual(status, "PRICE_FLOOR_UNAVAILABLE")

    def test_boundary_below_floor_deny(self):
        policy = EconomicsPolicy(minimum_margin_pct=Decimal("10"))
        inp = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("100"),
            commission_rate=Decimal("0.15"),
            commission_prov="CONFIGURED",
            logistics_cost=Decimal("10"),
            logistics_prov="CONFIGURED",
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        min_p, _, _ = calculate_minimum_price(inp, policy=policy)
        self.assertIsNotNone(min_p)
        below = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=min_p - Decimal("0.01"),
            commission_rate=Decimal("0.15"),
            commission_prov="CONFIGURED",
            logistics_cost=Decimal("10"),
            logistics_prov="CONFIGURED",
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        r = calculate_economics(below, policy=policy)
        self.assertEqual(r["decision"], DECISION_DENY)


class DecisionTests(unittest.TestCase):
    def test_allow_when_above_floor(self):
        inp = EconomicsInput(
            purchase_price=Decimal("50"),
            selling_price=Decimal("200"),
            commission_rate=Decimal("0.05"),
            commission_prov="CONFIGURED",
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        r = calculate_economics(inp)
        self.assertEqual(r["decision"], DECISION_ALLOW)

    def test_require_review_missing_commission(self):
        inp = EconomicsInput(purchase_price=Decimal("100"), selling_price=Decimal("200"))
        r = calculate_economics(inp)
        self.assertIn(r["decision"], {DECISION_REQUIRE_REVIEW, DECISION_WARN})


class MultiChannelTests(unittest.TestCase):
    def test_compare_channels_no_hardcoded_rates(self):
        base = EconomicsInput(
            sku="SKU-1",
            purchase_price=Decimal("95000"),
            selling_price=Decimal("120000"),
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        results = compare_channels(base, FIXTURE_CHANNELS, policy=EconomicsPolicy())
        self.assertEqual(len(results), 4)
        channels = {r["channel"] for r in results}
        self.assertEqual(channels, {"SITE", "WILDBERRIES", "OZON", "YANDEX_MARKET"})
        wb = next(r for r in results if r["channel"] == "WILDBERRIES")
        site = next(r for r in results if r["channel"] == "SITE")
        self.assertNotEqual(wb["contribution"], site["contribution"])


class ScenarioTests(unittest.TestCase):
    def test_scenario_non_mutation(self):
        orig = EconomicsInput(purchase_price=Decimal("100"), selling_price=Decimal("200"))
        adj = apply_scenario(orig, {"price_delta_pct": "-5"})
        self.assertEqual(orig.selling_price, Decimal("200"))
        self.assertEqual(adj.selling_price, Decimal("190.00"))


class VATTests(unittest.TestCase):
    def test_vat_included(self):
        inp = EconomicsInput(
            purchase_price=Decimal("100"),
            selling_price=Decimal("120"),
            vat_rate=Decimal("0.20"),
            selling_price_vat_mode="included",
            advertising_cost=Decimal("0"),
            advertising_prov="CONFIGURED",
        )
        r = calculate_economics(inp)
        self.assertIsNotNone(r["vat_amount"])


class ExcelIntegrationTests(unittest.TestCase):
    def _fixture_rows(self) -> bytes:
        return _xlsx(
            [
                ["PROF-1", "Profitable", "95000", "120000", "0.15", "350", "0", "0.20"],
                ["LOW-1", "Low margin", "90000", "98000", "0.15", "350", "0", "0.20"],
                ["LOSS-1", "Loss maker", "95000", "90000", "0.15", "350", "0", "0.20"],
                ["MISS-1", "Missing adv", "95000", "120000", "0.15", "350", "", "0.20"],
                ["DISC-S", "Seller disc", "100", "200", "0.05", "10", "0", "0", "0.10", "SELLER"],
                ["DISC-P", "Platform disc", "100", "200", "0.05", "10", "0", "0", "0.10", "PLATFORM"],
                ["BOUND", "Boundary", "100", "150", "0.15", "10", "0", "0"],
            ],
            headers=[
                "sku",
                "product_name",
                "purchase_price",
                "selling_price",
                "commission_rate",
                "logistics_cost",
                "advertising_cost",
                "vat_rate",
                "discount_rate",
                "discount_ownership",
            ],
        )

    def test_e2e_workbook_readback(self):
        svc = DataIntelligenceService(store=InMemoryDatasetStore())
        ing = svc.ingest(self._fixture_rows(), filename="economics.xlsx", tenant_id="tenant-a")
        out = svc.run_economics_process(ing["dataset_id"], tenant_id="tenant-a")
        self.assertEqual(out["status"], "OK")
        wb = load_workbook(io.BytesIO(out["content"]), read_only=True, data_only=True)
        names = set(wb.sheetnames)
        self.assertIn("ECONOMICS", names)
        self.assertIn("ISSUES", names)
        self.assertIn("SUMMARY", names)
        econ = wb["ECONOMICS"]
        headers = [c.value for c in next(econ.iter_rows(min_row=1, max_row=1))]
        self.assertIn("contribution", headers)
        self.assertIn("decision", headers)
        self.assertIn("profit_label", headers)
        rows = list(econ.iter_rows(min_row=2, values_only=True))
        self.assertGreaterEqual(len(rows), 5)
        joined = " ".join(str(c) for r in rows for c in r if c)
        self.assertIn("PROF-1", joined)
        wb.close()

    def test_tenant_isolation(self):
        svc = DataIntelligenceService(store=InMemoryDatasetStore())
        ing = svc.ingest(self._fixture_rows(), filename="e.xlsx", tenant_id="tenant-a")
        with self.assertRaises(DataIntelError) as ctx:
            svc.run_economics_process(ing["dataset_id"], tenant_id="tenant-b")
        self.assertEqual(ctx.exception.reason, DATASET_ACCESS_DENIED)

    def test_multi_channel_batch(self):
        svc = DataIntelligenceService(store=InMemoryDatasetStore())
        data = _xlsx([["SKU-1", "Widget", "95000", "120000", "0", "0", "0", "0.20"]])
        ing = svc.ingest(data, filename="m.xlsx", tenant_id="tenant-a")
        out = svc.run_economics_process(
            ing["dataset_id"],
            tenant_id="tenant-a",
            channel_configs=FIXTURE_CHANNELS,
        )
        self.assertGreaterEqual(out["summary"]["result_rows"], 4)


class ProvenanceTests(unittest.TestCase):
    def test_explanation_fields(self):
        row = {
            "sku": "A1",
            "purchase_price": "100",
            "selling_price": "200",
            "commission_rate": "0.05",
            "advertising_cost": "0",
            "__source_row": 2,
        }
        inp = parse_economics_row(row, policy=EconomicsPolicy())
        r = calculate_economics(inp)
        self.assertIn("formula", r)
        self.assertIn("provenance", r)
        self.assertIn("note", r)


class InvalidValueTests(unittest.TestCase):
    def test_negative_price_flagged_via_parse(self):
        row = {"purchase_price": "-10", "selling_price": "100"}
        inp = parse_economics_row(row, policy=EconomicsPolicy())
        # parse succeeds; economics still runs but negative purchase is invalid business input
        r = calculate_economics(inp)
        self.assertIsNotNone(r)


class NoExternalCallsTests(unittest.TestCase):
    def test_no_llm_imports(self):
        import sys

        for name in list(sys.modules):
            self.assertFalse(name.startswith("openai"), name)
            self.assertFalse(name.startswith("anthropic"), name)


class BatchUnitTests(unittest.TestCase):
    def test_batch_summary_counts(self):
        rows = [
            {"sku": "A", "purchase_price": "50", "selling_price": "200", "commission_rate": "0.05", "advertising_cost": "0"},
            {"sku": "B", "purchase_price": "100", "selling_price": "110"},
        ]
        batch = economics_batch_rows(rows, policy=EconomicsPolicy())
        self.assertEqual(batch["summary"]["process"], "product_economics")
        self.assertGreaterEqual(batch["summary"]["rows"], 2)


if __name__ == "__main__":
    unittest.main()

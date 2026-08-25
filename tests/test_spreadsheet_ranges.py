"""Unit tests for spreadsheet get_range after ingest."""

from __future__ import annotations

import io
import unittest

from openpyxl import Workbook

from documents.models import SOURCE_TEST_FIXTURE, DocumentIngestRequest
from documents.service import DocumentService
from documents.store import InMemoryDocumentStore
from memory.models import SCOPE_PROJECT, MemoryScope
from security.encryption import SENSITIVITY_INTERNAL


def _scope(sid="proj-range"):
    return MemoryScope(scope_type=SCOPE_PROJECT, scope_id=sid)


def _fixture():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B1"] = "y"
    ws["A2"] = 10
    ws["B2"] = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), ws.title


class SpreadsheetRangesTests(unittest.TestCase):
    def test_get_range_after_ingest(self):
        data, sheet_name = _fixture()
        svc = DocumentService(InMemoryDocumentStore())
        scope = _scope()
        row = svc.ingest(
            DocumentIngestRequest(
                scope=scope,
                filename="grid.xlsx",
                content=data,
                source_type=SOURCE_TEST_FIXTURE,
                source_id="rng-1",
                sensitivity=SENSITIVITY_INTERNAL,
            )
        )
        self.assertIn(row.document_id, svc._parsed_cache)
        rng, cells = svc.get_range(
            row.document_id,
            sheet_name=sheet_name,
            start_row=1,
            end_row=2,
            start_column=1,
            end_column=2,
            requesting_scope=scope,
        )
        self.assertEqual(rng.sheet_name, sheet_name)
        self.assertEqual(rng.cell_count, 4)
        self.assertEqual(rng.a1_range, "A1:B2")
        self.assertEqual(len(cells), 4)
        sheet = svc.get_sheet(row.document_id, sheet_name, requesting_scope=scope)
        self.assertEqual(sheet.sheet_name, sheet_name)


if __name__ == "__main__":
    unittest.main()

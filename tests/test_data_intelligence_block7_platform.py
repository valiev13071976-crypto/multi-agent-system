"""Block 7 Excel / Data Intelligence — platform closure tests."""

from __future__ import annotations

import io
import tempfile
import unittest
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_intel.errors import (
    DATASET_BATCH_REQUIRED,
    DATASET_CURRENCY_MISMATCH,
    DATASET_JOIN_EXPLOSION,
    DATASET_TYPE_MISMATCH,
    DataIntelError,
)
from data_intel.formulas import sanitize_cell_text
from data_intel.ingest import ingest_bytes
from data_intel.large import LargeDatasetPolicy, build_row_batch_plan
from data_intel.merge import merge_datasets
from data_intel.planner import (
    LARGE_BATCH_ROWS,
    LARGE_SYNC_ROWS,
    assert_hard_batch_admission,
    plan_data_job,
)
from data_intel.service import DataIntelligenceService
from data_intel.store import InMemoryDatasetStore
from documents.errors import ARCHIVE_EXPANSION_LIMIT_EXCEEDED
from documents.zip_safety import inspect_zip_safety
from side_effects.persistence import build_side_effect_persistence
from task_queue.lanes import LANE_BULK, LANE_INTERACTIVE, LaneCapacityConfig
from task_queue.queue import TaskQueue


def _csv_bytes(rows: list[list]) -> bytes:
    lines = []
    for r in rows:
        lines.append(",".join(str(c) for c in r))
    return "\n".join(lines).encode("utf-8")


def _xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for r in rows:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class PlannerAdmissionTests(unittest.TestCase):
    def test_large_batch_rows_requires_batch(self):
        planned = plan_data_job(
            dataset_id="ds-big",
            tenant_id="tenant-a",
            operations=("ingest",),
            row_count=LARGE_BATCH_ROWS,
        )
        self.assertTrue(planned.enqueue)
        self.assertEqual(planned.execution_lane, LANE_BULK)
        assert_hard_batch_admission(planned.trusted_metadata)

    def test_interactive_hint_cannot_downgrade(self):
        planned = plan_data_job(
            dataset_id="ds-h",
            tenant_id="tenant-a",
            operations=("merge",),
            row_count=LARGE_SYNC_ROWS,
            force_interactive_hint=True,
        )
        self.assertEqual(planned.execution_lane, LANE_BULK)

    def test_sync_allowed_below_threshold(self):
        planned = plan_data_job(
            dataset_id="ds-s",
            tenant_id="tenant-a",
            operations=("ingest",),
            row_count=LARGE_SYNC_ROWS - 1,
        )
        self.assertFalse(planned.enqueue)


class IngestSafetyTests(unittest.TestCase):
    def test_fake_xlsx_type_mismatch(self):
        with self.assertRaises(DataIntelError) as ctx:
            ingest_bytes(b"not a zip", filename="bad.xlsx", tenant_id="t1")
        self.assertEqual(ctx.exception.reason, DATASET_TYPE_MISMATCH)

    def test_xlsx_zip_bomb_bound(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/workbook.xml", b"0" * 2_000_000)
        with self.assertRaises(DataIntelError) as ctx:
            ingest_bytes(buf.getvalue(), filename="bomb.xlsx", tenant_id="t1")
        self.assertEqual(ctx.exception.reason, ARCHIVE_EXPANSION_LIMIT_EXCEEDED)


class BusinessLogicTests(unittest.TestCase):
    def test_currency_mismatch_not_compared(self):
        svc = DataIntelligenceService(store=InMemoryDatasetStore())
        result = svc.compare_prices(
            [{"sku": "A", "price": "10", "currency": "USD"}],
            [{"sku": "A", "price": "10", "currency": "EUR"}],
        )
        self.assertTrue(result["unresolved"])
        self.assertEqual(result["unresolved"][0]["reason"], DATASET_CURRENCY_MISMATCH)

    def test_many_to_many_join_rejected(self):
        left = [{"k": "x", "v": 1}, {"k": "x", "v": 2}]
        right = [{"k": "x", "v": 3}, {"k": "x", "v": 4}]
        with self.assertRaises(DataIntelError) as ctx:
            merge_datasets(left, right, keys=["k"], how="inner")
        self.assertEqual(ctx.exception.reason, DATASET_JOIN_EXPLOSION)

    def test_formula_injection_sanitized(self):
        self.assertTrue(sanitize_cell_text("=cmd|'/c calc'!A0").startswith("'"))


class LargeData300kTests(unittest.TestCase):
    def test_300k_classified_batch(self):
        planned = plan_data_job(
            dataset_id="ds-300k",
            tenant_id="tenant-a",
            operations=("ingest", "normalize"),
            row_count=300_000,
        )
        self.assertTrue(planned.enqueue)
        self.assertEqual(planned.trusted_metadata["trusted_job_type"], "data_large")

    def test_300k_batch_plan_bounded(self):
        plan = build_row_batch_plan(
            dataset_id="ds-300k",
            tenant_id="tenant-a",
            row_count=300_000,
            rows_per_batch=500,
        )
        self.assertEqual(plan["batch_count"], 600)
        self.assertTrue(all(b["bounded"] for b in plan["batches"]))

    def test_300k_sync_gate_fail_closed(self):
        svc = DataIntelligenceService(
            store=InMemoryDatasetStore(),
            large_policy=LargeDatasetPolicy(max_sync_rows=LARGE_SYNC_ROWS),
        )
        # Generate CSV header + 300k data rows in chunks without holding all in memory at once
        # For gate test use planner row_count directly on generate path
        data = _csv_bytes([["sku", "price"]] + [[f"S{i}", str(i)] for i in range(LARGE_BATCH_ROWS)])
        with self.assertRaises(DataIntelError) as ctx:
            svc.ingest(data, filename="big.csv", tenant_id="tenant-a", enqueue_large=False)
        # ingest at 300k without workflow should fail or async - with enqueue_large=False and no workflow:
        self.assertIn(ctx.exception.reason, {DATASET_BATCH_REQUIRED, "large_dataset_workflow_unavailable"})


class InteractiveIsolationTests(unittest.TestCase):
    def test_batch_flood_interactive_runnable(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            path = str(Path(tmp) / "data-flood.sqlite3")
            bundle = build_side_effect_persistence(
                env={
                    "SIDE_EFFECT_PERSISTENCE_BACKEND": "sqlite",
                    "SIDE_EFFECT_DB_PATH": path,
                    "SIDE_EFFECT_RECOVERY_SCAN_ON_STARTUP": "false",
                    "MAX_RUNNING_GLOBAL": "10",
                    "INTERACTIVE_RESERVED": "3",
                },
                durable=True,
                run_recovery_scan=False,
            )
            q = TaskQueue(
                store=bundle.task_queue_store,
                lease_seconds=60,
                lane_config=LaneCapacityConfig(
                    interactive_reserved=3, background_may_borrow=False
                ),
            )
            for i in range(25):
                planned = plan_data_job(
                    dataset_id=f"ds-{i}",
                    tenant_id="tenant-bulk",
                    operations=("analyze",),
                    row_count=LARGE_SYNC_ROWS,
                )
                q.enqueue(
                    workflow_id=f"data-{i}",
                    task_id=f"t-{i}",
                    execution_key=f"data-flood-{i}",
                    tenant_id="tenant-bulk",
                    execution_lane=planned.execution_lane,
                    priority="low",
                    metadata=dict(planned.trusted_metadata),
                )
            for i in range(8):
                t = q.dequeue(
                    worker_id=f"wb{i}",
                    max_running_global=10,
                    max_running_per_tenant=50,
                )
                if t is None:
                    break
            q.enqueue(
                workflow_id="ix",
                task_id="tix",
                execution_key="ek-ix-data",
                tenant_id="tenant-ix",
                execution_lane=LANE_INTERACTIVE,
                priority="high",
            )
            claimed = q.dequeue(
                worker_id="w-ix",
                max_running_global=10,
                max_running_per_tenant=50,
            )
            self.assertIsNotNone(claimed)
            self.assertEqual(claimed.execution_lane, LANE_INTERACTIVE)


class TenantIsolationTests(unittest.TestCase):
    def test_cross_tenant_profile_denied(self):
        svc = DataIntelligenceService(store=InMemoryDatasetStore())
        data = _xlsx_bytes({"S": [["sku", "price"], ["A", "1"]]})
        result = svc.ingest(data, filename="p.xlsx", tenant_id="tenant-a")
        with self.assertRaises(DataIntelError) as ctx:
            svc.profile(result["dataset_id"], tenant_id="tenant-b")
        self.assertEqual(ctx.exception.reason, "dataset_access_denied")


class GenerationE2ETests(unittest.TestCase):
    def test_generate_and_reingest(self):
        svc = DataIntelligenceService(store=InMemoryDatasetStore())
        data = _xlsx_bytes(
            {
                "Data": [
                    ["sku", "price"],
                    ["A-1", "10.5"],
                    ["A-2", "20"],
                ]
            }
        )
        ing = svc.ingest(data, filename="src.xlsx", tenant_id="tenant-a")
        gen = svc.generate_excel(ing["dataset_id"], tenant_id="tenant-a")
        self.assertGreater(gen["size"], 0)
        re = ingest_bytes(gen["content"], filename="out.xlsx", tenant_id="tenant-a")
        self.assertGreaterEqual(re.descriptor.row_count, 2)


if __name__ == "__main__":
    unittest.main()

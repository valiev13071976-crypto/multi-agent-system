"""Unit tests: spreadsheet formulas stored as data, never executed."""

from __future__ import annotations

import io
import unittest

from openpyxl import Workbook

from documents.models import CELL_FORMULA, SOURCE_TEST_FIXTURE, DocumentIngestRequest
from documents.service import DocumentService
from documents.store import InMemoryDocumentStore
from memory.models import SCOPE_PROJECT, MemoryScope
from security.encryption import SENSITIVITY_INTERNAL


def _scope(sid="proj-formula"):
    return MemoryScope(scope_type=SCOPE_PROJECT, scope_id=sid)


class SpreadsheetFormulaSafetyTests(unittest.TestCase):
    def test_formula_not_executed(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["A2"] = 2
        ws["A3"] = "=A1+A2"
        buf = io.BytesIO()
        wb.save(buf)

        svc = DocumentService(InMemoryDocumentStore())
        row = svc.ingest(
            DocumentIngestRequest(
                scope=_scope(),
                filename="calc.xlsx",
                content=buf.getvalue(),
                source_type=SOURCE_TEST_FIXTURE,
                source_id="xlsx-1",
                sensitivity=SENSITIVITY_INTERNAL,
            )
        )
        parsed = svc._parsed_cache[row.document_id]
        formulas = [c for c in parsed.cells if c.value_type == CELL_FORMULA]
        self.assertTrue(formulas)
        self.assertTrue(all(c.formula and c.formula.startswith("=") for c in formulas))
        self.assertFalse(any(c.value in {"3", 3} for c in formulas))
        self.assertFalse(any(c.cached_value for c in formulas))


if __name__ == "__main__":
    unittest.main()

"""Finished business XLSX generation — searchable, formatted, no macros."""

from __future__ import annotations

import io
from datetime import datetime, timezone

from data_intel.errors import EXCEL_GENERATION_FAILED, DataIntelError
from data_intel.formulas import sanitize_cell_text
from data_intel.identifiers_ru import normalize_inn


def _col_letter(idx: int) -> str:
    n = idx
    letters = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters or "A"


def _write_table(ws, headers: list[str], rows: list[list], *, text_cols: set[int] | None = None):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    text_cols = text_cols or set()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, sanitize_cell_text(h))
        cell.font = Font(bold=True)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            if c - 1 in text_cols:
                ws.cell(r, c, sanitize_cell_text(val)).number_format = "@"
            else:
                ws.cell(r, c, sanitize_cell_text(val) if isinstance(val, str) else val)
    # widths
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(40, max(12, len(str(headers[c - 1])) + 4))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_col_letter(len(headers))}{max(1, len(rows) + 1)}"
    if rows and headers:
        try:
            tab = Table(displayName=f"T{ws.title[:20].replace(' ', '')}{ws.max_row}", ref=ws.auto_filter.ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)
        except Exception:
            pass


def generate_workbook(
    *,
    sheets: dict[str, dict],
    provenance: dict | None = None,
    summary: dict | None = None,
) -> bytes:
    """
    sheets: {name: {"headers": [...], "rows": [[...]], "text_cols": [idx...]}}
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise DataIntelError(EXCEL_GENERATION_FAILED) from exc
    try:
        wb = Workbook()
        # remove default
        default = wb.active
        wb.remove(default)

        if summary is not None:
            ws = wb.create_sheet("SUMMARY", 0)
            from openpyxl.styles import Font

            bold = Font(bold=True)
            ws["A1"] = "Metric"
            ws["B1"] = "Value"
            ws["A1"].font = bold
            ws["B1"].font = bold
            r = 2
            for k, v in summary.items():
                ws.cell(r, 1, sanitize_cell_text(k))
                ws.cell(r, 2, sanitize_cell_text(v))
                r += 1
            ws.freeze_panes = "A2"

        for name, payload in sheets.items():
            safe_name = str(name)[:31] or "Sheet"
            ws = wb.create_sheet(safe_name)
            headers = list(payload.get("headers") or [])
            rows = list(payload.get("rows") or [])
            text_cols = set(payload.get("text_cols") or ())
            _write_table(ws, headers, rows, text_cols=text_cols)

        # Provenance / source sheet
        ws = wb.create_sheet("Provenance")
        prov = dict(provenance or {})
        prov.setdefault("generated_at", datetime.now(timezone.utc).isoformat())
        ws["A1"] = "Key"
        ws["B1"] = "Value"
        r = 2
        for k, v in prov.items():
            ws.cell(r, 1, sanitize_cell_text(k))
            ws.cell(r, 2, sanitize_cell_text(v))
            r += 1

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        _validate_xlsx_bytes(data)
        return data
    except DataIntelError:
        raise
    except Exception as exc:
        raise DataIntelError(EXCEL_GENERATION_FAILED) from exc


def _validate_xlsx_bytes(data: bytes) -> None:
    if not data or len(data) < 4:
        raise DataIntelError(EXCEL_GENERATION_FAILED)
    if data[:2] != b"PK":
        raise DataIntelError(EXCEL_GENERATION_FAILED)
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True)
        if not wb.sheetnames:
            raise DataIntelError(EXCEL_GENERATION_FAILED)
        wb.close()
    except DataIntelError:
        raise
    except Exception as exc:
        raise DataIntelError(EXCEL_GENERATION_FAILED) from exc


def generate_business_result_workbook(
    *,
    result_headers: list[str],
    result_rows: list[list],
    issue_headers: list[str],
    issue_rows: list[list],
    summary: dict,
    provenance: dict | None = None,
    text_cols: set[int] | None = None,
) -> bytes:
    """Canonical Block 10 output: RESULT + ISSUES + SUMMARY (+ Provenance)."""
    return generate_workbook(
        summary=dict(summary or {}),
        sheets={
            "RESULT": {
                "headers": list(result_headers),
                "rows": list(result_rows),
                "text_cols": set(text_cols or ()),
            },
            "ISSUES": {
                "headers": list(issue_headers),
                "rows": list(issue_rows),
                "text_cols": set(),
            },
        },
        provenance=dict(provenance or {}),
    )


def generate_searchable_payments_workbook(rows: list[dict], *, title: str = "Payments") -> bytes:
    """Business workbook: INN/company as text, filters, all related rows retained."""
    headers = [
        "inn",
        "company_name",
        "amount",
        "payment_date",
        "document_number",
        "purpose",
        "row_ref",
    ]
    out_rows = []
    for r in rows:
        inn = normalize_inn(r.get("inn"))
        inn_val = inn.normalized if inn.normalized else (r.get("inn") or "")
        out_rows.append(
            [
                inn_val,
                r.get("company_name") or r.get("counterparty") or r.get("name") or "",
                r.get("amount") or "",
                r.get("payment_date") or r.get("date") or "",
                r.get("document_number") or "",
                r.get("purpose") or "",
                r.get("__row_ref") or r.get("row_ref") or "",
            ]
        )
    # Index sheet for quick INN lookup list
    inns = sorted({str(r[0]) for r in out_rows if r[0]})
    return generate_workbook(
        summary={"title": title, "rows": len(out_rows), "distinct_inn": len(inns)},
        sheets={
            "Data": {"headers": headers, "rows": out_rows, "text_cols": {0, 1, 4}},
            "INN_Index": {
                "headers": ["inn"],
                "rows": [[i] for i in inns],
                "text_cols": {0},
            },
        },
        provenance={"kind": "searchable_payments", "title": title},
    )


def generate_comparison_workbook(comparison: dict) -> bytes:
    def _flat(items, kind):
        rows = []
        for it in items:
            if kind in {"matched", "changed"}:
                rows.append(
                    [
                        it.get("key"),
                        it.get("sku"),
                        it.get("ean"),
                        it.get("product"),
                        it.get("old_price"),
                        it.get("new_price"),
                        it.get("abs_diff"),
                        it.get("pct_diff"),
                    ]
                )
            elif kind in {"new", "missing"}:
                row = it.get("row") or {}
                rows.append(
                    [
                        it.get("key"),
                        row.get("sku"),
                        row.get("ean"),
                        row.get("product_name") or row.get("name"),
                        row.get("price"),
                        it.get("supplier"),
                    ]
                )
            else:
                rows.append([str(it.get("key")), str(it.get("match") or it)])
        return rows

    sheets = {
        "Matched": {
            "headers": ["key", "sku", "ean", "product", "old_price", "new_price", "abs_diff", "pct_diff"],
            "rows": _flat(comparison.get("matched") or [], "matched"),
            "text_cols": {0, 1, 2},
        },
        "Changed": {
            "headers": ["key", "sku", "ean", "product", "old_price", "new_price", "abs_diff", "pct_diff"],
            "rows": _flat(comparison.get("changed") or [], "changed"),
            "text_cols": {0, 1, 2},
        },
        "New": {
            "headers": ["key", "sku", "ean", "product", "price", "supplier"],
            "rows": _flat(comparison.get("new") or [], "new"),
            "text_cols": {0, 1, 2},
        },
        "Missing": {
            "headers": ["key", "sku", "ean", "product", "price", "supplier"],
            "rows": _flat(comparison.get("missing") or [], "missing"),
            "text_cols": {0, 1, 2},
        },
        "Conflicts": {
            "headers": ["key", "detail"],
            "rows": _flat(comparison.get("conflicts") or [], "conflicts"),
            "text_cols": {0},
        },
        "Unresolved": {
            "headers": ["key", "detail"],
            "rows": _flat(comparison.get("unresolved") or [], "unresolved"),
            "text_cols": {0},
        },
    }
    return generate_workbook(
        summary=dict(comparison.get("summary") or {}),
        sheets=sheets,
        provenance={"kind": "price_comparison"},
    )

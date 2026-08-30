"""XLS / XLSX / CSV ingestion into grid + structure (reuses Documents parsers where possible)."""

from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass

from data_intel.contracts import DatasetDescriptor, TableDescriptor, new_id, utc_now
from data_intel.cleaning import clean_text
from data_intel.errors import (
    DATASET_PARSE_FAILED,
    DATASET_TYPE_MISMATCH,
    UNSUPPORTED_SPREADSHEET,
    DataIntelError,
)
from data_intel.structure import detect_tables_in_sheet, extract_table_rows


@dataclass
class SheetGrid:
    name: str
    rows: list[list]
    hidden: bool = False


@dataclass
class IngestResult:
    descriptor: DatasetDescriptor
    sheets: list[SheetGrid]
    table_rows: dict[str, list[dict]]  # table_id -> rows (with __source_row)


def _checksum(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _detect_format(filename: str, data: bytes) -> str:
    name = (filename or "").lower()
    magic_xlsx = data[:2] == b"PK"
    magic_xls = data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    if name.endswith(".xlsx"):
        if magic_xlsx:
            return "xlsx"
        if magic_xls:
            raise DataIntelError(DATASET_TYPE_MISMATCH)
        raise DataIntelError(DATASET_TYPE_MISMATCH)
    if name.endswith(".xls"):
        if magic_xls:
            return "xls"
        if magic_xlsx:
            raise DataIntelError(DATASET_TYPE_MISMATCH)
        raise DataIntelError(DATASET_TYPE_MISMATCH)
    if magic_xlsx:
        return "xlsx"
    if magic_xls:
        return "xls"
    if name.endswith(".csv") or b"," in data[:200] or b";" in data[:200]:
        return "csv"
    raise DataIntelError(UNSUPPORTED_SPREADSHEET)


def _decode_csv(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = data.decode(enc)
            # Strip leftover BOM character if present
            if text.startswith("\ufeff"):
                text = text.lstrip("\ufeff")
            return text
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace").lstrip("\ufeff")


def load_csv_sheets(data: bytes) -> list[SheetGrid]:
    text = _decode_csv(data)
    sample = text[:4096]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    if sample.count("\t") > max(sample.count(";"), sample.count(",")):
        delimiter = "\t"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [list(r) for r in reader]
    # Drop completely empty trailing rows
    while rows and all(clean_text(c) is None for c in rows[-1]):
        rows.pop()
    return [SheetGrid(name="Sheet1", rows=rows)]


def load_xlsx_sheets(data: bytes) -> list[SheetGrid]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataIntelError(DATASET_PARSE_FAILED) from exc
    if data[:2] == b"PK":
        from documents.errors import DocumentError as DocError
        from documents.zip_safety import inspect_zip_safety

        try:
            inspect_zip_safety(data)
        except DocError as exc:
            raise DataIntelError(str(exc.reason)) from exc
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    except Exception as exc:
        raise DataIntelError(DATASET_PARSE_FAILED) from exc
    sheets: list[SheetGrid] = []
    try:
        for ws in wb.worksheets:
            hidden = getattr(ws, "sheet_state", "visible") != "visible"
            rows: list[list] = []
            for row in ws.iter_rows(values_only=False):
                cells = []
                for cell in row:
                    if getattr(cell, "data_type", None) == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        # Keep formula text as displayed placeholder; never evaluate
                        cells.append(str(cell.value) if cell.value else None)
                    else:
                        cells.append(cell.value)
                rows.append(cells)
            sheets.append(SheetGrid(name=ws.title, rows=rows, hidden=hidden))
    finally:
        wb.close()
    return sheets


def load_xls_sheets(data: bytes) -> list[SheetGrid]:
    try:
        import xlrd
    except ImportError as exc:
        raise DataIntelError(UNSUPPORTED_SPREADSHEET) from exc
    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise DataIntelError(DATASET_PARSE_FAILED) from exc
    sheets = []
    for i in range(book.nsheets):
        sh = book.sheet_by_index(i)
        rows = []
        for r in range(sh.nrows):
            rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
        sheets.append(SheetGrid(name=sh.name, rows=rows))
    return sheets


def ingest_bytes(
    data: bytes,
    *,
    filename: str,
    tenant_id: str,
    source_document_id: str = "",
    dataset_id: str | None = None,
) -> IngestResult:
    if not data:
        raise DataIntelError(DATASET_PARSE_FAILED)
    fmt = _detect_format(filename, data)
    try:
        if fmt == "csv":
            sheets = load_csv_sheets(data)
        elif fmt == "xlsx":
            sheets = load_xlsx_sheets(data)
        elif fmt == "xls":
            sheets = load_xls_sheets(data)
        else:
            raise DataIntelError(UNSUPPORTED_SPREADSHEET)
    except DataIntelError:
        raise
    except Exception as exc:
        raise DataIntelError(DATASET_PARSE_FAILED) from exc

    ds_id = dataset_id or new_id("ds-")
    tables: list[TableDescriptor] = []
    table_rows: dict[str, list[dict]] = {}
    for sheet in sheets:
        detected = detect_tables_in_sheet(sheet.name, sheet.rows, dataset_id=ds_id)
        for t in detected:
            tables.append(t)
            table_rows[t.table_id] = extract_table_rows(sheet.rows, t)

    row_count = sum(len(v) for v in table_rows.values())
    col_count = max((len(t.columns) for t in tables), default=0)
    desc = DatasetDescriptor(
        dataset_id=ds_id,
        tenant_id=tenant_id,
        source_document_id=source_document_id or "",
        format=fmt,
        sheets=tuple(s.name for s in sheets),
        tables=tuple(tables),
        row_count=row_count,
        column_count=col_count,
        checksum=_checksum(data),
        created_at=utc_now(),
        provenance={
            "filename": filename,
            "hidden_sheets": [s.name for s in sheets if s.hidden],
            "ingestion": "data_intel.ingest",
        },
    )
    return IngestResult(descriptor=desc, sheets=sheets, table_rows=table_rows)

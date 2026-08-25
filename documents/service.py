"""Canonical DocumentService — business logic owner."""

from __future__ import annotations

import re
import uuid
from pathlib import Path

from autonomy.models import sanitize_metadata
from documents.access import (
    OP_DELETE,
    OP_EXTRACT,
    OP_INGEST,
    OP_READ,
    DocumentAccessDenied,
    DocumentAccessPolicy,
)
from documents.chunker import DocumentChunker
from documents.errors import (
    DOCUMENT_DISABLED,
    DOCUMENT_ENCRYPTION_UNAVAILABLE,
    DOCUMENT_PATH_DENIED,
    DOCUMENT_RANGE_INVALID,
    DOCUMENT_SECRET_DENIED,
    DOCUMENT_SHEET_NOT_FOUND,
    DOCUMENT_STORE_UNAVAILABLE,
    DOCUMENT_TOO_LARGE,
    DocumentError,
)
from documents.models import (
    CellRange,
    DocumentChunkRecord,
    DocumentIngestRequest,
    DocumentProvenance,
    DocumentRecord,
    DocumentSearchRequest,
    DocumentSearchResult,
    STATUS_DELETED,
    STATUS_FAILED,
    STATUS_INGESTED,
    STATUS_PARSED,
    STATUS_PARTIAL,
    citation_ref_for,
    content_hash_bytes,
    sanitize_filename,
)
from documents.parsers import DocumentParserRegistry, build_default_registry
from documents.store import DocumentStore, _clone
from documents.type_detect import resolve_document_type
from documents.validator import DocumentValidator
from memory.models import MEMORY_SEMANTIC, SOURCE_DOCUMENT, MemoryIngestRequest, MemoryScope, utc_now
from security.encryption import (
    ENCRYPTION_REQUIRED,
    SENSITIVITY_INTERNAL,
    SENSITIVITY_SECRET,
    EncryptionService,
    EncryptionUnavailableError,
)
from security.redaction import redact


_SECRET_MARKERS = (
    "GITHUB_WRITE_TOKEN",
    "PANDA_ENCRYPTION_KEY",
    "sk-",
    "ghp_",
    "Bearer ",
    "Authorization:",
)


class DocumentService:
    def __init__(
        self,
        store: DocumentStore,
        *,
        registry: DocumentParserRegistry | None = None,
        access: DocumentAccessPolicy | None = None,
        chunker: DocumentChunker | None = None,
        validator: DocumentValidator | None = None,
        encryption: EncryptionService | None = None,
        memory_service=None,
        observability=None,
        limits: dict | None = None,
        allowed_roots: tuple[str, ...] = (),
        enabled: bool = True,
    ):
        self.store = store
        self.registry = registry or build_default_registry(
            max_file_bytes=int((limits or {}).get("max_file_bytes", 5_000_000))
        )
        self.access = access or DocumentAccessPolicy()
        self.chunker = chunker or DocumentChunker()
        self.validator = validator or DocumentValidator()
        self.encryption = encryption
        self.memory_service = memory_service
        self.observability = observability
        self.limits = {
            "max_file_bytes": 5_000_000,
            "max_text_bytes": 1_000_000,
            "max_table_cells": 100_000,
            "max_sheets": 50,
            "max_pages": 200,
            "max_chunks": 500,
            **dict(limits or {}),
        }
        self.allowed_roots = tuple(str(Path(r).resolve()) for r in allowed_roots)
        self.enabled = bool(enabled)
        self.blocked_reason: str | None = None
        self._parsed_cache: dict[str, object] = {}

    def ingest(
        self,
        request: DocumentIngestRequest,
        *,
        requesting_scope: MemoryScope | None = None,
    ) -> DocumentRecord:
        if not self.enabled or self.blocked_reason:
            raise DocumentError(self.blocked_reason or DOCUMENT_DISABLED)
        if not getattr(self.store, "available", True):
            raise DocumentError(DOCUMENT_STORE_UNAVAILABLE)
        req_scope = requesting_scope or request.scope
        self._require_access(req_scope, request.scope, OP_INGEST, request.filename)

        data = bytes(request.content)
        if len(data) > int(self.limits["max_file_bytes"]):
            self._emit("document.failed", status="failed", metadata={"reason": DOCUMENT_TOO_LARGE})
            raise DocumentError(DOCUMENT_TOO_LARGE)
        if request.sensitivity == SENSITIVITY_SECRET or self._looks_like_secret(data):
            self._emit("document.denied", status="denied", metadata={"reason": DOCUMENT_SECRET_DENIED})
            self._metric("document_ingest_total", "unknown", "denied", "secret")
            raise DocumentError(DOCUMENT_SECRET_DENIED)

        digest = content_hash_bytes(data)
        existing = self.store.find_by_hash(request.scope, digest)
        if existing is not None and existing.status not in {STATUS_DELETED, STATUS_FAILED}:
            self._emit("document.deduplicated", status=existing.status, metadata={"document_type": existing.document_type})
            self._metric("document_dedup_total", existing.document_type, "ok", existing.sensitivity)
            return existing

        try:
            doc_type, media = resolve_document_type(
                filename=request.filename,
                data=data,
                declared_media_type=request.media_type,
            )
        except DocumentError as exc:
            self._emit("document.failed", status="failed", metadata={"reason": exc.reason})
            self._metric("document_parse_failure_total", "unknown", "failed", request.sensitivity)
            raise

        stamp = utc_now()
        document_id = str(uuid.uuid4())
        provenance = DocumentProvenance(
            source_type=request.source_type,
            source_id=request.source_id,
            ingested_by=request.ingested_by,
            ingested_at=stamp,
            source_hash=digest,
            workflow_id=request.workflow_id,
            task_id=request.task_id,
        )
        record = DocumentRecord(
            document_id=document_id,
            scope=request.scope,
            filename_safe=sanitize_filename(request.filename),
            media_type=media,
            document_type=doc_type,
            size_bytes=len(data),
            content_hash=digest,
            source_type=request.source_type,
            source_ref=request.source_id,
            provenance=provenance,
            sensitivity=request.sensitivity,
            status=STATUS_INGESTED,
            created_at=stamp,
            updated_at=stamp,
            metadata_safe=dict(request.metadata_safe),
        )
        created = self.store.create(record, provenance, tags=request.tags)
        if created.document_id != document_id:
            self._emit("document.deduplicated", status=created.status, metadata={"document_type": created.document_type})
            return created

        self._emit("document.ingested", status="ingested", metadata={"document_type": doc_type})
        self._metric("document_ingest_total", doc_type, "ingested", request.sensitivity)
        self._metric("document_bytes_total", doc_type, "ingested", request.sensitivity, amount=len(data))

        # Parse immediately for foundation completeness
        try:
            parsed_record = self._parse_and_persist(created, data)
        except DocumentError as exc:
            failed = _clone(created, status=STATUS_FAILED, updated_at=utc_now())
            try:
                self.store.update(failed, expected_version=created.version)
            except Exception:
                pass
            self._emit("document.failed", status="failed", metadata={"reason": exc.reason, "document_type": doc_type})
            self._metric("document_parse_failure_total", doc_type, "failed", request.sensitivity)
            raise

        if request.promote_to_memory and self.memory_service is not None:
            self._promote_chunks_to_memory(parsed_record.document_id, requesting_scope=req_scope)
        return parsed_record

    def ingest_trusted_path(
        self,
        path: str | Path,
        *,
        scope: MemoryScope,
        source_type: str,
        source_id: str,
        sensitivity: str = "internal",
        requesting_scope: MemoryScope | None = None,
    ) -> DocumentRecord:
        resolved = self._resolve_trusted_path(path)
        data = resolved.read_bytes()
        return self.ingest(
            DocumentIngestRequest(
                scope=scope,
                filename=resolved.name,
                content=data,
                source_type=source_type,
                source_id=source_id,
                sensitivity=sensitivity,
            ),
            requesting_scope=requesting_scope,
        )

    def get(self, document_id: str, *, requesting_scope: MemoryScope) -> DocumentRecord | None:
        row = self.store.get(document_id)
        if row is None or row.status == STATUS_DELETED:
            return None
        try:
            self.access.require(requesting=requesting_scope, target=row.scope, operation=OP_READ)
        except DocumentAccessDenied:
            return None
        return row

    def delete(self, document_id: str, *, requesting_scope: MemoryScope, reason: str = "delete") -> DocumentRecord:
        row = self.store.get(document_id)
        if row is None:
            raise DocumentError("document_not_found")
        if row.status == STATUS_DELETED:
            return row
        self._require_access(requesting_scope, row.scope, OP_DELETE, row.document_type)
        deleted = self.store.delete(document_id, expected_version=row.version)
        self._parsed_cache.pop(document_id, None)
        self._emit("document.deleted", status="deleted", metadata={"reason": reason, "document_type": row.document_type})
        return deleted

    def list_chunks(self, document_id: str, *, requesting_scope: MemoryScope) -> tuple[DocumentChunkRecord, ...]:
        row = self.get(document_id, requesting_scope=requesting_scope)
        if row is None:
            return ()
        self.access.require(requesting=requesting_scope, target=row.scope, operation=OP_EXTRACT)
        return self.store.list_chunks(document_id)

    def get_sheet(self, document_id: str, sheet_name: str, *, requesting_scope: MemoryScope):
        parsed = self._require_parsed(document_id, requesting_scope)
        for sheet in parsed.sheets:
            if sheet.sheet_name == sheet_name:
                self._emit("spreadsheet.inspected", status="ok", metadata={"document_type": "xlsx"})
                self._metric("spreadsheet_sheet_total", "xlsx", "ok", "internal")
                return sheet
        raise DocumentError(DOCUMENT_SHEET_NOT_FOUND)

    def get_range(
        self,
        document_id: str,
        *,
        sheet_name: str,
        start_row: int,
        end_row: int,
        start_column: int,
        end_column: int,
        requesting_scope: MemoryScope,
    ) -> tuple[CellRange, tuple]:
        parsed = self._require_parsed(document_id, requesting_scope)
        if end_row < start_row or end_column < start_column:
            raise DocumentError(DOCUMENT_RANGE_INVALID)
        if not any(s.sheet_name == sheet_name for s in parsed.sheets):
            raise DocumentError(DOCUMENT_SHEET_NOT_FOUND)
        cell_count = (end_row - start_row + 1) * (end_column - start_column + 1)
        if cell_count > int(self.limits["max_table_cells"]):
            raise DocumentError(DOCUMENT_RANGE_INVALID)
        from openpyxl.utils import get_column_letter

        a1 = (
            f"{get_column_letter(start_column)}{start_row}:"
            f"{get_column_letter(end_column)}{end_row}"
        )
        rng = CellRange(
            sheet_name=sheet_name,
            start_row=start_row,
            end_row=end_row,
            start_column=start_column,
            end_column=end_column,
            a1_range=a1,
            cell_count=cell_count,
        )
        selected = tuple(
            c
            for c in parsed.cells
            if c.metadata_safe.get("sheet_name") == sheet_name
            and start_row <= c.row <= end_row
            and start_column <= c.column <= end_column
        )
        self._emit("spreadsheet.range_extracted", status="ok", metadata={"document_type": "xlsx"})
        self._metric("spreadsheet_range_extract_total", "xlsx", "ok", "internal")
        return rng, selected

    def search(self, request: DocumentSearchRequest, *, requesting_scope: MemoryScope | None = None):
        req = requesting_scope or request.scope
        self._require_access(req, request.scope, OP_READ, "search")
        docs = self.store.list_by_scope(request.scope)
        if request.document_types:
            allowed = set(request.document_types)
            docs = tuple(d for d in docs if d.document_type in allowed)
        q = request.query.lower().strip()
        results = []
        for doc in docs:
            for chunk in self.store.list_chunks(doc.document_id):
                text = chunk.content_safe or ""
                if chunk.encrypted_content and self.encryption is not None and not text:
                    try:
                        text = self.encryption.decrypt(chunk.encrypted_content)
                    except Exception:
                        text = ""
                if q and q not in text.lower():
                    continue
                score = 1.0 if not q else (text.lower().count(q) / max(1, len(text.split())))
                results.append(
                    DocumentSearchResult(
                        document_id=doc.document_id,
                        chunk_id=chunk.chunk_id,
                        score=round(float(score), 6),
                        snippet_safe=(text[:240]),
                        source_location=chunk.source_location,
                        provenance=dict(chunk.provenance_json),
                        citation_ref=citation_ref_for(doc.document_id, chunk.chunk_id),
                    )
                )
        results.sort(key=lambda r: (-r.score, r.document_id, r.chunk_id))
        return tuple(results[: request.limit])

    def _parse_and_persist(self, record: DocumentRecord, data: bytes) -> DocumentRecord:
        parser = self.registry.get_parser(record.document_type)
        try:
            parsed = parser.parse(
                document_id=record.document_id,
                data=data,
                filename=record.filename_safe,
                limits=self.limits,
            )
        except DocumentError:
            raise
        except Exception as exc:
            raise DocumentError("document_parse_failed") from exc

        self.validator.validate_parsed(parsed, limits=self.limits)
        provenance_payload = {
            "document_id": record.document_id,
            "source_type": record.source_type,
            "source_id": record.source_ref,
            "parser_id": parsed.parser_id,
            "parser_version": parsed.parser_version,
        }
        chunks = self.chunker.chunk(
            parsed,
            scope=record.scope,
            sensitivity=record.sensitivity,
            provenance=provenance_payload,
        )
        secured = []
        for ch in chunks:
            content = ch.content_safe or ""
            content_safe = redact(content)
            encrypted = None
            if record.sensitivity in ENCRYPTION_REQUIRED:
                if self.encryption is None:
                    raise DocumentError(DOCUMENT_ENCRYPTION_UNAVAILABLE)
                try:
                    encrypted = self.encryption.encrypt(content_safe).serialize()
                except EncryptionUnavailableError as exc:
                    raise DocumentError(DOCUMENT_ENCRYPTION_UNAVAILABLE) from exc
                content_safe = None
            secured.append(
                DocumentChunkRecord(
                    chunk_id=ch.chunk_id,
                    document_id=ch.document_id,
                    scope=ch.scope,
                    ordinal=ch.ordinal,
                    content_hash=ch.content_hash,
                    source_location=ch.source_location,
                    content_safe=content_safe,
                    encrypted_content=encrypted,
                    sensitivity=record.sensitivity,
                    provenance_json=dict(ch.provenance_json),
                    metadata_safe=dict(ch.metadata_safe),
                    created_at=ch.created_at,
                )
            )
        self.store.save_chunks(record.document_id, tuple(secured))
        status = STATUS_PARTIAL if parsed.partial or parsed.warnings else STATUS_PARSED
        updated = _clone(
            record,
            status=status,
            page_count=parsed.pages,
            sheet_count=parsed.workbook.sheet_count if parsed.workbook else None,
            chunk_count=len(secured),
            parser_version=parsed.parser_version,
            title=parsed.title or record.title,
            warnings=parsed.warnings,
            updated_at=utc_now(),
            metadata_safe={**dict(record.metadata_safe), **dict(parsed.metadata_safe)},
        )
        saved = self.store.update(updated, expected_version=record.version)
        self._parsed_cache[record.document_id] = parsed
        self._emit(
            "document.partial" if status == STATUS_PARTIAL else "document.parsed",
            status=status,
            metadata={"document_type": record.document_type, "parser_id": parsed.parser_id},
        )
        self._emit("document.chunked", status=status, metadata={"document_type": record.document_type})
        self._metric("document_parse_total", record.document_type, status, record.sensitivity)
        self._metric("document_chunk_total", record.document_type, status, record.sensitivity, amount=len(secured))
        if parsed.workbook:
            self._metric("spreadsheet_sheet_total", "xlsx", status, record.sensitivity, amount=parsed.workbook.sheet_count)
        return saved

    def _promote_chunks_to_memory(self, document_id: str, *, requesting_scope: MemoryScope) -> None:
        chunks = self.store.list_chunks(document_id)
        for ch in chunks:
            text = ch.content_safe
            if not text and ch.encrypted_content and self.encryption is not None:
                try:
                    text = self.encryption.decrypt(ch.encrypted_content)
                except Exception:
                    continue
            if not text:
                continue
            try:
                self.memory_service.ingest(
                    MemoryIngestRequest(
                        scope=ch.scope,
                        memory_type=MEMORY_SEMANTIC,
                        content=text,
                        source_type=SOURCE_DOCUMENT,
                        source_id=document_id,
                        sensitivity=ch.sensitivity if ch.sensitivity != SENSITIVITY_SECRET else SENSITIVITY_INTERNAL,
                        confidence=0.6,
                        tags=("document",),
                        created_by_component="document_service",
                        external_reference=ch.chunk_id,
                        metadata_safe={
                            "document_id": document_id,
                            "chunk_id": ch.chunk_id,
                            "source_location": ch.source_location,
                            "content_hash": ch.content_hash,
                        },
                    ),
                    requesting_scope=requesting_scope,
                    validated=False,
                    auto=False,
                )
            except Exception:
                # Memory write policy may deny; document ingest still succeeds.
                continue

    def _require_parsed(self, document_id: str, requesting_scope: MemoryScope):
        row = self.get(document_id, requesting_scope=requesting_scope)
        if row is None:
            raise DocumentAccessDenied()
        self.access.require(requesting=requesting_scope, target=row.scope, operation=OP_EXTRACT)
        cached = self._parsed_cache.get(document_id)
        if cached is not None:
            return cached
        raise DocumentError("document_parse_failed")

    def _resolve_trusted_path(self, path: str | Path) -> Path:
        raw = str(path or "")
        if not raw or "\x00" in raw:
            raise DocumentError(DOCUMENT_PATH_DENIED)
        candidate = Path(raw)
        # Reject obvious traversal tokens before resolve
        parts = candidate.as_posix().replace("\\", "/").split("/")
        if ".." in parts:
            raise DocumentError(DOCUMENT_PATH_DENIED)
        try:
            resolved = candidate.resolve(strict=False)
        except Exception as exc:
            raise DocumentError(DOCUMENT_PATH_DENIED) from exc
        if not self.allowed_roots:
            raise DocumentError(DOCUMENT_PATH_DENIED)
        ok = False
        for root in self.allowed_roots:
            try:
                resolved.relative_to(root)
                ok = True
                break
            except ValueError:
                continue
        if not ok or not resolved.is_file():
            raise DocumentError(DOCUMENT_PATH_DENIED)
        return resolved

    def _require_access(self, requesting, target, operation, hint: str = "") -> None:
        try:
            self.access.require(requesting=requesting, target=target, operation=operation)
        except DocumentAccessDenied:
            self._emit(
                "document.denied",
                status="denied",
                metadata={
                    "reason_code": "document_scope_access_denied",
                    "operation": operation,
                    "scope_type": requesting.scope_type,
                    "target_scope_type": target.scope_type,
                },
            )
            self._metric("document_ingest_total" if operation == OP_INGEST else "document_parse_total", "unknown", "denied", "internal")
            raise

    def _looks_like_secret(self, data: bytes) -> bool:
        try:
            sample = data[:4096].decode("utf-8", errors="ignore")
        except Exception:
            return False
        if any(m in sample for m in _SECRET_MARKERS):
            return True
        return bool(re.search(r"(?i)\b(api[_-]?key|password|secret)\b\s*[:=]", sample))

    def _emit(self, event_type: str, *, status: str = "", metadata: dict | None = None) -> None:
        obs = self.observability
        if obs is None:
            return
        try:
            safe = sanitize_metadata(
                {
                    k: v
                    for k, v in dict(metadata or {}).items()
                    if k not in {"content", "query", "scope_id", "filename", "path"}
                }
            )
            obs.emit(event_type, component="documents", status=status, metadata=safe)
        except Exception:
            pass

    def _metric(
        self,
        name: str,
        document_type: str,
        status: str,
        sensitivity: str,
        *,
        amount: int = 1,
        parser_id: str = "n_a",
    ) -> None:
        obs = self.observability
        if obs is None or not getattr(obs, "metrics", None):
            return
        try:
            obs.metrics.inc(
                name,
                labels={
                    "component": "documents",
                    "document_type": document_type or "unknown",
                    "parser_id": parser_id[:64],
                    "status": status,
                    "sensitivity": sensitivity or "unknown",
                },
                amount=amount,
            )
        except Exception:
            pass

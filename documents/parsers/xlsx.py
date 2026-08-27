"""XLSX parser — read-only, formulas as data, no macros/external fetch."""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from decimal import Decimal

from documents.errors import (
    DOCUMENT_MACROS_NOT_ALLOWED,
    DOCUMENT_PARSE_FAILED,
    DOCUMENT_TOO_MANY_CELLS,
    DOCUMENT_TOO_MANY_SHEETS,
    DocumentError,
)
from documents.models import (
    CELL_BLANK,
    CELL_BOOLEAN,
    CELL_DATE,
    CELL_DATETIME,
    CELL_FORMULA,
    CELL_NUMBER,
    CELL_STRING,
    DOC_XLSX,
    CellValue,
    ParsedDocument,
    TableBlock,
    TextBlock,
    WorkbookRecord,
    WorksheetRecord,
    content_hash_text,
)
from documents.zip_safety import inspect_zip_safety


def _col_letter(idx: int) -> str:
    n = idx
    letters = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters or "A"


def _cell_type_and_value(cell) -> tuple[str, str | None, str | None, bool]:
    """Return value_type, value, formula, cached_flag. Never evaluates formulas."""
    if cell.value is None and (getattr(cell, "data_type", None) != "f"):
        return CELL_BLANK, None, None, False
    # openpyxl: data_type 'f' means formula; value may be formula string when data_only=False
    if getattr(cell, "data_type", None) == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    ):
        formula = str(cell.value) if cell.value is not None else None
        # With data_only=False, there is no calculated cache unless workbook had cached values
        # exposed separately; treat as formula-only data.
        return CELL_FORMULA, None, formula, False
    val = cell.value
    if isinstance(val, bool):
        return CELL_BOOLEAN, "true" if val else "false", None, False
    if isinstance(val, datetime):
        return CELL_DATETIME, val.isoformat(), None, False
    if isinstance(val, date):
        return CELL_DATE, val.isoformat(), None, False
    if isinstance(val, (int, float, Decimal)):
        # Avoid float loss for large ints
        if isinstance(val, float) and not val.is_integer():
            return CELL_NUMBER, format(Decimal(str(val)), "f"), None, False
        if isinstance(val, float) and val.is_integer():
            return CELL_NUMBER, str(int(val)), None, False
        return CELL_NUMBER, str(val), None, False
    text = str(val)
    potential = text[:1] in {"=", "+", "-", "@"}
    return CELL_STRING, text, None, False if not potential else False


class XlsxDocumentParser:
    parser_id = "xlsx_v1"
    version = "1.0.0"
    supported_types = (DOC_XLSX,)

    def parse(self, *, document_id: str, data: bytes, filename: str, limits: dict) -> ParsedDocument:
        max_sheets = int(limits.get("max_sheets", 50))
        max_cells = int(limits.get("max_table_cells", 100_000))
        zip_info = inspect_zip_safety(data)
        names_l = " ".join(zip_info["names"]).lower()
        if "vbaProject.bin".lower() in names_l or filename.lower().endswith(".xlsm"):
            raise DocumentError(DOCUMENT_MACROS_NOT_ALLOWED)
        has_external = any(
            "externalLink" in n or "externallinks" in n.lower() for n in zip_info["names"]
        )
        warnings = []
        if has_external:
            warnings.append("external_links_present")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DocumentError(DOCUMENT_PARSE_FAILED) from exc

        try:
            wb = load_workbook(
                filename=io.BytesIO(data),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise DocumentError(DOCUMENT_PARSE_FAILED) from exc

        try:
            sheet_names = tuple(wb.sheetnames)
            if len(sheet_names) > max_sheets:
                raise DocumentError(DOCUMENT_TOO_MANY_SHEETS)
            sheets = []
            cells = []
            tables = []
            text_blocks = []
            cell_count = 0
            for idx, name in enumerate(sheet_names):
                ws = wb[name]
                max_row = int(ws.max_row or 0)
                max_col = int(ws.max_column or 0)
                sheets.append(
                    WorksheetRecord(
                        sheet_name=name,
                        index=idx,
                        max_row=max_row,
                        max_column=max_col,
                        visible=True,
                        merged_ranges_count=0,
                    )
                )
                # Bounded cell sample for structure
                rows_data = []
                headers = []
                for r_i, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=min(max_row, 200), max_col=min(max_col, 50)),
                    start=1,
                ):
                    row_vals = []
                    for cell in row:
                        cell_count += 1
                        if cell_count > max_cells:
                            raise DocumentError(DOCUMENT_TOO_MANY_CELLS)
                        # openpyxl may yield EmptyCell without coordinate for sparse formulas
                        if not hasattr(cell, "coordinate"):
                            row_vals.append("")
                            continue
                        vtype, value, formula, cached = _cell_type_and_value(cell)
                        coord = cell.coordinate
                        potential = bool(
                            value and str(value)[:1] in {"=", "+", "-", "@"} and vtype == CELL_STRING
                        )
                        cells.append(
                            CellValue(
                                row=cell.row,
                                column=cell.column,
                                coordinate=coord,
                                value=value,
                                value_type=vtype,
                                formula=formula,
                                display_value=value if cached else (formula if formula else value),
                                cached_value=cached,
                                potential_formula_text=potential,
                                metadata_safe={"sheet_name": name},
                            )
                        )
                        row_vals.append(formula or value or "")
                    if r_i == 1:
                        headers = tuple(row_vals) or tuple(
                            _col_letter(i + 1) for i in range(len(row_vals))
                        )
                    else:
                        rows_data.append(tuple(row_vals))
                tables.append(
                    TableBlock(
                        table_id=str(uuid.uuid4()),
                        ordinal=idx,
                        name=name,
                        rows=tuple(rows_data[:100]),
                        columns=headers,
                        source_location=f"xlsx:sheet:{name}",
                        metadata_safe={"max_row": max_row, "max_column": max_col},
                    )
                )
                preview_lines = [f"Sheet: {name}"]
                if headers:
                    preview_lines.append(" | ".join(headers))
                for row in rows_data[:10]:
                    preview_lines.append(" | ".join(row))
                preview = "\n".join(preview_lines)
                text_blocks.append(
                    TextBlock(
                        block_id=str(uuid.uuid4()),
                        ordinal=idx,
                        text=preview,
                        content_hash=content_hash_text(preview),
                        source_location=f"xlsx:sheet:{name}:preview",
                        section=name,
                    )
                )
            workbook = WorkbookRecord(
                document_id=document_id,
                sheet_names=sheet_names,
                sheet_count=len(sheet_names),
                active_sheet=sheet_names[0] if sheet_names else None,
                has_macros=False,
                has_external_links=has_external,
                metadata_safe={"cell_sample_count": len(cells)},
            )
            return ParsedDocument(
                document_id=document_id,
                text_blocks=tuple(text_blocks),
                tables=tuple(tables),
                metadata_safe={
                    "filename": filename,
                    "sheet_count": len(sheet_names),
                    "has_external_links": has_external,
                },
                parser_id=self.parser_id,
                parser_version=self.version,
                title=filename,
                sheets=tuple(sheets),
                workbook=workbook,
                cells=tuple(cells),
                warnings=tuple(warnings),
            )
        finally:
            try:
                wb.close()
            except Exception:
                pass
